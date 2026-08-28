"""Offline verification for Day-4d export functionality.

Runs the real generators against realistic payloads for all eight tasks. No
Qdrant, no LLM, no Office: every assertion reads the produced file's own bytes,
so this is runnable anywhere.

    python tests/acceptance/verify_exports.py [--keep]

--keep leaves the generated files in tests/acceptance/_export_output/ so they
can be opened by hand in Word / Excel / Adobe Reader for the live check.
"""

from __future__ import annotations

import argparse
import io
import re
import sys
import threading
import time
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import load_workbook  # noqa: E402
from pypdf import PdfReader  # noqa: E402

from matter_clerk import export, pleadings  # noqa: E402
from matter_clerk.citation import Citation  # noqa: E402
from matter_clerk.export.payload import ExportPayload  # noqa: E402
from matter_clerk.structured import (  # noqa: E402
    ComparisonRow,
    ComparisonTable,
    EntityCategory,
    EntityRow,
    TimelineRow,
)

OUT_DIR = Path(__file__).resolve().parent / "_export_output"

_passed = 0
_failed: list[str] = []


def check(label: str, condition: bool, detail: str = "") -> None:
    global _passed
    if condition:
        _passed += 1
        print(f"  [ok]   {label}")
    else:
        _failed.append(f"{label}{(' - ' + detail) if detail else ''}")
        print(f"  [FAIL] {label}{(' - ' + detail) if detail else ''}")


# --------------------------------------------------------------------------
# Fixtures — the Imperial Plaza matter
# --------------------------------------------------------------------------
CITATIONS = [
    Citation(
        source="Imperial_Plaza_Lease.pdf",
        page_or_paragraph="p.4",
        text_snippet="The Manager shall remit all rents collected to the Owner within ten (10) business days of receipt.",
    ),
    Citation(
        source="Cresthaven_Correspondence.eml",
        page_or_paragraph="from Kevin Oskoui, 2024-05-11",
        text_snippet="We have not received the April or May remittance. Please confirm status by return.",
    ),
    Citation(
        source="Cresthaven_Management_Agreement.pdf",
        page_or_paragraph="p.9 (OCR)",
        text_snippet="Either party may terminate this agreement on sixty (60) days written notice.",
    ),
]

BASE = dict(
    citations=CITATIONS,
    matter_name="Imperial Plaza v. Cresthaven",
    source_files=[
        "Imperial_Plaza_Lease.pdf",
        "Cresthaven_Correspondence.eml",
        "Cresthaven_Management_Agreement.pdf",
    ],
    model="xiaomi/mimo-v2.5-pro",
    embed_model="BAAI/bge-small-en-v1.5",
    top_k=14,
    timestamp="2026-08-11T14:22:05+00:00",
)

PLEADING_BODY = """\
**STATEMENT OF CLAIM**

1. The Plaintiff, Imperial Plaza Holdings Inc., is a corporation incorporated
under the laws of Ontario. [Imperial_Plaza_Lease.pdf p.4]

2. By a written management agreement, the Plaintiff retained the Defendant to
manage the property. [Cresthaven_Management_Agreement.pdf p.9 (OCR)]

7. [ELEMENTS REQUIRED — lawyer to complete: name the cause of action and plead
each element. The retrieved passages do not establish the standard of care.]

8. The Defendant failed to remit rents for April and May 2024.
[Cresthaven_Correspondence.eml from Kevin Oskoui, 2024-05-11]

9. [ADDITIONAL MATERIAL REQUIRED — lawyer to complete: state the quantum of
damages. No passage quantifies the shortfall.]
"""

TIMELINE_MD = """\
| Date | Event | Source | Procedural significance |
|---|---|---|---|
| 2023-03-14 | Management agreement executed | [Imperial_Plaza_Lease.pdf p.4] | |
| March 2024 | Remittances become irregular | [Cresthaven_Management_Agreement.pdf p.9 (OCR)] | |
| 2024-05-11 | Owner demands confirmation of status | [Cresthaven_Correspondence.eml from Kevin Oskoui, 2024-05-11] | Starts the 60-day notice period |
"""

COMPARISON_MD = """\
| Attribute | Imperial_Plaza_Lease.pdf | Cresthaven_Management_Agreement.pdf |
|---|---|---|
| Clause located at | s. 12.1 [Imperial_Plaza_Lease.pdf p.4] | Not present in this document |
| Notice period | Ten business days [Imperial_Plaza_Lease.pdf p.4] | - |
"""


def payload_for(task: str) -> ExportPayload:
    if task == "draft_pleading":
        return ExportPayload(
            task=task, task_label="Draft Pleading", answer_markdown=PLEADING_BODY,
            is_pleading=True, party_role="Plaintiff",
            pleading_warnings=["The matter's files do not contain typical pleading markers."],
            request_summary={"Pleading type": pleadings.SOC},
            provenance_label="Drew on", **BASE,
        )
    if task == "timeline":
        return ExportPayload(
            task=task, task_label="Timeline", answer_markdown=TIMELINE_MD,
            timeline_rows=[
                TimelineRow(date="2023-03-14", date_iso="2023-03-14",
                            event="Management agreement executed",
                            source="[Imperial_Plaza_Lease.pdf p.4]"),
                TimelineRow(date="March 2024", event="Remittances become irregular",
                            source="[Cresthaven_Management_Agreement.pdf p.9 (OCR)]"),
                TimelineRow(date="2024-05-11", date_iso="2024-05-11",
                            event="Owner demands confirmation of status",
                            source="[Cresthaven_Correspondence.eml from Kevin Oskoui, 2024-05-11]",
                            significance="Starts the 60-day notice period"),
            ],
            provenance_label="Drew on", **BASE,
        )
    if task == "find_entities":
        return ExportPayload(
            task=task, task_label="Find Entities",
            answer_markdown="### Parties\n\n| Entity | Count | Source |\n|---|---|---|\n| Imperial Plaza Holdings Inc. | 4 | [Imperial_Plaza_Lease.pdf p.4] |\n",
            entity_categories=[
                EntityCategory(name="Parties", rows=[
                    EntityRow(entity="Imperial Plaza Holdings Inc.", count=4,
                              source="[Imperial_Plaza_Lease.pdf p.4]"),
                    EntityRow(entity="Cresthaven Property Management Ltd.", count=6,
                              source="[Cresthaven_Management_Agreement.pdf p.9 (OCR)]"),
                ]),
                EntityCategory(name="Dollar amounts", rows=[
                    EntityRow(entity="$48,250.00", count=1,
                              source="[Cresthaven_Correspondence.eml from Kevin Oskoui, 2024-05-11]"),
                ]),
                EntityCategory(name="Addresses", rows=[],
                               note="None found in the retrieved passages."),
            ],
            provenance_label="Drew on", **BASE,
        )
    if task == "compare_clauses":
        return ExportPayload(
            task=task, task_label="Compare Clauses", answer_markdown=COMPARISON_MD,
            comparison_table=ComparisonTable(
                files=["Imperial_Plaza_Lease.pdf", "Cresthaven_Management_Agreement.pdf"],
                rows=[
                    ComparisonRow(attribute="Clause located at",
                                  cells=["s. 12.1 [Imperial_Plaza_Lease.pdf p.4]",
                                         "Not present in this document"]),
                    ComparisonRow(attribute="Notice period",
                                  cells=["Ten business days [Imperial_Plaza_Lease.pdf p.4]", "-"]),
                ],
            ),
            provenance_label="Compared across", **BASE,
        )
    labels = {
        "summarize": "Summarize", "find_facts": "Find Facts",
        "draft_memo": "Draft Memo", "draft_correspondence": "Draft Correspondence",
    }
    return ExportPayload(
        task=task, task_label=labels[task],
        answer_markdown=(
            "## Issue\n\nWhether the Defendant breached the remittance obligation.\n\n"
            "## Facts\n\nThe agreement requires remittance within ten business days. "
            "[Imperial_Plaza_Lease.pdf p.4]\n\n"
            "- April remittance not received\n- May remittance not received\n"
        ),
        provenance_label="Drew on", **BASE,
    )


ALL_TASKS = [
    "summarize", "find_facts", "draft_memo", "draft_correspondence",
    "timeline", "find_entities", "compare_clauses", "draft_pleading",
]


# --------------------------------------------------------------------------
# Readers
# --------------------------------------------------------------------------
def docx_xml(data: bytes) -> tuple[str, str]:
    """(document.xml, all footer xml concatenated)."""
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        doc = z.read("word/document.xml").decode("utf-8")
        footers = "".join(
            z.read(n).decode("utf-8")
            for n in z.namelist()
            if re.match(r"word/footer\d*\.xml", n)
        )
    return doc, footers


def docx_text(data: bytes) -> str:
    doc, footers = docx_xml(data)
    return re.sub(r"<[^>]+>", "", doc + footers)


def pdf_text(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


# --------------------------------------------------------------------------
# Suites
# --------------------------------------------------------------------------
def verify_all_formats(keep: bool) -> None:
    print("\n== Every applicable format for every task ==")
    for task in ALL_TASKS:
        payload = payload_for(task)
        formats = export.list_export_formats(payload)
        expect = ["docx", "pdf"] + (["xlsx"] if task in export.EXCEL_TASKS else [])
        check(f"{task}: offers {expect}", formats == expect, str(formats))

        for fmt in formats:
            data, mimetype, filename = export.generate(payload, fmt)
            magic = {
                "docx": b"PK", "xlsx": b"PK", "pdf": b"%PDF",
            }[fmt]
            check(
                f"{task}.{fmt}: valid file ({len(data):,} bytes)",
                data.startswith(magic) and len(data) > 1000,
                f"magic={data[:4]!r}",
            )
            if keep:
                OUT_DIR.mkdir(parents=True, exist_ok=True)
                (OUT_DIR / filename).write_bytes(data)


def verify_metadata_footer() -> None:
    print("\n== Attribution footer present in every export ==")
    for task in ALL_TASKS:
        payload = payload_for(task)
        expected_date = "2026-08-11"
        for fmt in export.list_export_formats(payload):
            data, _, _ = export.generate(payload, fmt)
            if fmt == "docx":
                text = docx_text(data)
            elif fmt == "pdf":
                text = pdf_text(data)
            else:
                wb = load_workbook(io.BytesIO(data))
                text = " ".join(
                    str(c.value)
                    for ws in wb.worksheets
                    for row in ws.iter_rows()
                    for c in row
                    if c.value is not None
                )
            ok = "Generated by Matter Clerk" in text and expected_date in text
            has_model = "mimo" in text
            check(f"{task}.{fmt}: attribution + model", ok and has_model)


def verify_long_file_list() -> None:
    """Regression: a matter with many files must not blow the 255-char OOXML
    property limit.

    The live Day-4d bug: the full attribution (which names every source file)
    was assigned to the docx `comments` property, and python-docx raised
    "exceeded 255 char limit for property" on any matter with more than two or
    three real filenames. The split is short-attribution-in-properties,
    full-attribution-in-footer, so this asserts both halves.
    """
    print("\n== Long file lists / 255-char property limit ==")
    many_files = [
        "Imperial_Plaza_Master_Lease_Agreement_Executed_2023.pdf",
        "Cresthaven_Property_Management_Agreement_Amended.pdf",
        "Cresthaven_Correspondence_Oskoui_May_2024.eml",
        "Imperial_Plaza_Rent_Roll_and_Remittance_Ledger_2024.xlsx",
        "Cresthaven_Termination_Notice_Draft_Unsigned.docx",
        "Imperial_Plaza_Board_Minutes_2024_03_14_scanned.pdf",
        "Cresthaven_Bank_Statements_April_May_2024_redacted.pdf",
    ]
    over = BASE | {
        "source_files": many_files,
        # A matter name long enough to overflow the property field on its own,
        # so the clamp is exercised too and not just the file list.
        "matter_name": "Imperial Plaza Holdings Inc. v. Cresthaven Property "
        "Management Ltd. and 1274855 Ontario Inc. carrying on business as "
        "Cresthaven Realty Advisory Group " * 3,
    }
    payload = ExportPayload(
        task="draft_pleading", task_label="Draft Pleading",
        answer_markdown=PLEADING_BODY, is_pleading=True, party_role="Plaintiff",
        request_summary={"Pleading type": pleadings.SOC},
        provenance_label="Drew on", **over,
    )
    prose = ExportPayload(
        task="summarize", task_label="Summarize",
        answer_markdown="The Defendant failed to remit. [Imperial_Plaza_Lease.pdf p.4]",
        provenance_label="Drew on", **over,
    )
    tabular = payload_for("timeline").model_copy(update=over)

    check("attribution names all 7 files",
          all(f in prose.attribution() for f in many_files))
    check("full attribution really is over the limit (the bug's precondition)",
          len(prose.attribution()) > 255, str(len(prose.attribution())))
    check("short attribution fits the property limit",
          len(prose.short_attribution()) <= 255, str(len(prose.short_attribution())))
    check("short attribution names the tool and the date",
          prose.short_attribution().startswith("Matter Clerk - ")
          and prose.short_attribution().endswith("2026-08-11"),
          prose.short_attribution())

    for label, p in (("pleading", payload), ("prose", prose), ("timeline", tabular)):
        # 1. Word must generate at all — this is the call that used to raise.
        try:
            data, _, _ = export.generate(p, "docx")
            check(f"{label}.docx: generates with 7 files", data.startswith(b"PK"))
        except Exception as e:  # noqa: BLE001
            check(f"{label}.docx: generates with 7 files", False, str(e))
            continue

        with zipfile.ZipFile(io.BytesIO(data)) as z:
            core = z.read("docProps/core.xml").decode("utf-8")
        values = re.findall(r"<[^>]+>([^<]+)</[^>]+>", core)
        longest = max((len(v) for v in values), default=0)
        check(f"{label}.docx: every core property within 255 chars",
              longest <= 255, f"longest={longest}")
        check(f"{label}.docx: property holds the short attribution",
              "Matter Clerk - Imperial Plaza Holdings" in core)
        check(f"{label}.docx: property does NOT carry the file list",
              "Generated by Matter Clerk on" not in core)

        # 2. The full attribution moved to the body footer, intact.
        _, footer_xml = docx_xml(data)
        footer = re.sub(r"<[^>]+>", "", footer_xml)
        check(f"{label}.docx: footer carries the full attribution",
              "Generated by Matter Clerk on 2026-08-11" in footer
              and "Model: xiaomi/mimo-v2.5-pro" in footer)
        check(f"{label}.docx: footer names every one of the 7 files",
              all(f in footer for f in many_files))

        # 3. PDF: same split, and the drawn footer must not truncate away the
        #    model name the way the old 3-line cut did.
        pdata, _, _ = export.generate(p, "pdf")
        ptext = pdf_text(pdata)
        check(f"{label}.pdf: generates with 7 files", pdata.startswith(b"%PDF"))
        check(f"{label}.pdf: footer keeps date AND model",
              "Generated by Matter Clerk on 2026-08-11" in ptext
              and "Model: xiaomi/mimo-v2.5-pro" in ptext)
        check(f"{label}.pdf: footer lists files, eliding the tail if it must",
              all(f in ptext for f in many_files) or "more file" in ptext)
        subject = (PdfReader(io.BytesIO(pdata)).metadata or {}).get("/Subject", "")
        check(f"{label}.pdf: metadata subject is short, not the file list",
              "Generated by Matter Clerk on" not in str(subject), str(subject)[:80])

    # 4. Excel is unaffected — a cell has no property limit — but assert the
    #    full attribution still lands there so the split did not leak into it.
    xdata, _, _ = export.generate(tabular, "xlsx")
    wb = load_workbook(io.BytesIO(xdata))
    xtext = " ".join(
        str(c.value) for ws in wb.worksheets for row in ws.iter_rows()
        for c in row if c.value is not None
    )
    check("timeline.xlsx: keeps the full attribution with all 7 files",
          "Generated by Matter Clerk on 2026-08-11" in xtext
          and all(f in xtext for f in many_files))


def verify_draft_machinery(keep: bool) -> None:
    print("\n== Draft Pleading DRAFT machinery (safety-critical) ==")
    payload = payload_for("draft_pleading")

    docx_data, _, _ = export.generate(payload, "docx")
    doc_xml, footer_xml = docx_xml(docx_data)
    text = docx_text(docx_data)

    banner_hits = text.count(pleadings.DRAFT_BANNER)
    check("docx: DRAFT banner appears at head AND foot", banner_hits >= 2, f"count={banner_hits}")
    check("docx: banner drawn on dark-red shading", 'w:fill="7A0000"' in doc_xml)
    check("docx: banner also in the page footer", pleadings.DRAFT_BANNER in footer_xml)
    check("docx: cover note intro present",
          "This document is a DRAFT produced by an automated tool" in text)
    for phrase in ("Not reviewed by counsel", "Limitation periods and deadlines",
                   "admission consequences", "Do not file or serve"):
        check(f"docx: cover note point '{phrase[:28]}'", phrase in text)
    check("docx: cover note box has a red border", 'w:color="7A0000"' in doc_xml)
    check("docx: gap markers highlighted", 'w:highlight w:val="yellow"' in doc_xml)
    check("docx: ELEMENTS REQUIRED marker present", "ELEMENTS REQUIRED" in text)
    check("docx: ADDITIONAL MATERIAL REQUIRED marker present",
          "ADDITIONAL MATERIAL REQUIRED" in text)
    with zipfile.ZipFile(io.BytesIO(docx_data)) as z:
        core = z.read("docProps/core.xml").decode("utf-8")
    check("docx: document properties assert DRAFT", "NOT FOR FILING" in core)
    check("docx: pleading paragraph numbers preserved (7, 8, 9)",
          all(f"{n}." in text for n in (7, 8, 9)))

    pdf_data, _, _ = export.generate(payload, "pdf")
    ptext = pdf_text(pdf_data)
    check("pdf: DRAFT banner at head and foot", ptext.count(pleadings.DRAFT_BANNER) >= 2)
    check("pdf: watermark drawn on every page", ptext.count("NOT FOR FILING") >= 2)
    check("pdf: cover note present",
          "This document is a DRAFT produced by an automated tool" in ptext)
    check("pdf: gap markers present",
          "ELEMENTS REQUIRED" in ptext and "ADDITIONAL MATERIAL REQUIRED" in ptext)
    check("pdf: paragraph numbers preserved", "7." in ptext and "9." in ptext)

    # A pleading download must announce itself as a draft in the filename.
    for fmt in ("docx", "pdf"):
        _, _, name = export.generate(payload, fmt)
        check(f"{fmt}: filename is DRAFT-prefixed", name.startswith("DRAFT-"), name)


def verify_citation_fidelity() -> None:
    print("\n== Citation text survives export byte-exact ==")
    for task in ("draft_pleading", "summarize", "timeline"):
        payload = payload_for(task)
        for fmt in ("docx", "pdf"):
            data, _, _ = export.generate(payload, fmt)
            text = docx_text(data) if fmt == "docx" else pdf_text(data)
            flat = re.sub(r"\s+", "", text)
            check(
                f"{task}.{fmt}: underscored filename intact",
                "Imperial_Plaza_Lease.pdf" in flat,
                "underscores were eaten (markdown emphasis bug)",
            )
        # OCR marker must survive — it tells the lawyer the snippet is approximate.
        data, _, _ = export.generate(payload, "pdf")
        check(f"{task}.pdf: OCR locator marker preserved", "(OCR)" in pdf_text(data))


def verify_excel_shape() -> None:
    print("\n== Excel structure ==")
    # --- Timeline
    data, _, _ = export.generate(payload_for("timeline"), "xlsx")
    wb = load_workbook(io.BytesIO(data))
    check("timeline: sheets", wb.sheetnames == ["Timeline", "Citations", "Metadata"],
          str(wb.sheetnames))
    ws = wb["Timeline"]
    headers = [ws.cell(5, c).value for c in range(1, 5)]
    check("timeline: headers on row 5",
          headers == ["Date", "Event", "Source", "Procedural significance"], str(headers))
    check("timeline: headers bold", ws.cell(5, 1).font.bold is True)
    check("timeline: metadata note at top", "Imperial Plaza" in str(ws.cell(1, 1).value))
    import datetime as _dt
    check("timeline: complete date is a real Excel date",
          isinstance(ws.cell(6, 1).value, (_dt.date, _dt.datetime)), repr(ws.cell(6, 1).value))
    check("timeline: partial date stays verbatim text",
          ws.cell(7, 1).value == "March 2024", repr(ws.cell(7, 1).value))
    check("timeline: columns auto-fitted", ws.column_dimensions["B"].width > 10)
    check("timeline: header row frozen", ws.freeze_panes == "A6", str(ws.freeze_panes))

    # --- Find Entities
    data, _, _ = export.generate(payload_for("find_entities"), "xlsx")
    wb = load_workbook(io.BytesIO(data))
    check("entities: one sheet per category (+cites/meta)",
          wb.sheetnames == ["Parties", "Dollar amounts", "Addresses", "Citations", "Metadata"],
          str(wb.sheetnames))
    check("entities: empty category states so",
          "None found" in str(wb["Addresses"].cell(5, 1).value))
    check("entities: dollar amount verbatim",
          wb["Dollar amounts"].cell(6, 1).value == "$48,250.00")

    # --- Compare Clauses
    data, _, _ = export.generate(payload_for("compare_clauses"), "xlsx")
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Comparison"]
    check("comparison: files as columns",
          [ws.cell(5, c).value for c in range(1, 4)]
          == ["Attribute", "Imperial_Plaza_Lease.pdf", "Cresthaven_Management_Agreement.pdf"])
    check("comparison: 'Clause located at' is the first row",
          ws.cell(6, 1).value == "Clause located at")
    check("comparison: citation preserved inline in the cell",
          "[Imperial_Plaza_Lease.pdf p.4]" in str(ws.cell(6, 2).value))
    meta = " ".join(
        f"{r[0].value}={r[1].value}" for r in wb["Metadata"].iter_rows(min_col=1, max_col=2)
        if r[0].value
    )
    check("comparison: metadata records absent files",
          "Cresthaven_Management_Agreement.pdf" in meta and "Marked not present" in meta)
    check("comparison: metadata records the data source",
          "structured task output" in meta)


def verify_excel_refused_for_prose() -> None:
    print("\n== Excel refused for prose tasks ==")
    for task in ("summarize", "find_facts", "draft_memo", "draft_correspondence",
                 "draft_pleading"):
        payload = payload_for(task)
        check(f"{task}: no Excel button",
              "xlsx" not in export.list_export_formats(payload))
        try:
            export.generate(payload, "xlsx")
            check(f"{task}: endpoint refuses xlsx", False, "no exception raised")
        except export.UnsupportedFormat:
            check(f"{task}: endpoint refuses xlsx", True)


def verify_markdown_fallback() -> None:
    print("\n== Fallback when structured data is absent ==")
    payload = payload_for("timeline").model_copy(update={"timeline_rows": None})
    data, _, _ = export.generate(payload, "xlsx")
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Timeline"]
    check("fallback: markdown table parsed into the sheet",
          ws.cell(6, 2).value == "Management agreement executed", repr(ws.cell(6, 2).value))
    meta = " ".join(str(c.value) for r in wb["Metadata"].iter_rows() for c in r)
    check("fallback: metadata says the export was degraded", "fallback" in meta)


def verify_cache() -> None:
    print("\n== Result cache ==")
    export.clear()
    payload = payload_for("timeline")
    token = export.store_result(payload)
    check("token is url-safe and long", len(token) >= 32 and "/" not in token)
    check("lookup returns the payload", export.get_result(token) is payload)
    check("SECOND lookup still returns it (no delete-on-read)",
          export.get_result(token) is payload)
    check("unknown token -> None", export.get_result("nope") is None)

    # Concurrency: two simultaneous exports of one token must both succeed.
    export.clear()
    token = export.store_result(payload_for("compare_clauses"))
    results: list[tuple[str, int]] = []
    errors: list[str] = []
    barrier = threading.Barrier(6)

    def worker(fmt: str) -> None:
        try:
            barrier.wait(timeout=10)
            p = export.get_result(token)
            data, _, _ = export.generate(p, fmt)
            results.append((fmt, len(data)))
        except Exception as e:  # noqa: BLE001
            errors.append(f"{fmt}: {e}")

    threads = [
        threading.Thread(target=worker, args=(f,))
        for f in ("docx", "pdf", "xlsx", "docx", "pdf", "xlsx")
    ]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=60)
    check("6 concurrent exports of one token all succeed",
          len(results) == 6 and not errors, "; ".join(errors))
    check("concurrent exports produced identical sizes per format",
          len({f for f, _ in results}) == 3
          and all(
              len({size for fmt, size in results if fmt == f}) == 1
              for f in ("docx", "pdf", "xlsx")
          ))

    # TTL eviction
    export.clear()
    import matter_clerk.export.cache as cache_mod

    original = cache_mod.TTL_SECONDS
    try:
        cache_mod.TTL_SECONDS = 1
        t2 = export.store_result(payload)
        check("fresh token resolves", export.get_result(t2) is not None)
        time.sleep(1.2)
        check("expired token is reaped", export.get_result(t2) is None)
    finally:
        cache_mod.TTL_SECONDS = original

    # LRU cap
    export.clear()
    for _ in range(cache_mod.MAX_ENTRIES + 25):
        export.store_result(payload)
    check(f"cache capped at {cache_mod.MAX_ENTRIES}",
          export.stats()["size"] == cache_mod.MAX_ENTRIES, str(export.stats()))
    export.clear()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--keep", action="store_true",
                    help="write generated files to _export_output/ for manual opening")
    args = ap.parse_args()

    print("Day-4d export verification")
    verify_all_formats(args.keep)
    verify_metadata_footer()
    verify_long_file_list()
    verify_draft_machinery(args.keep)
    verify_citation_fidelity()
    verify_excel_shape()
    verify_excel_refused_for_prose()
    verify_markdown_fallback()
    verify_cache()

    print(f"\n{_passed} passed, {len(_failed)} failed")
    if _failed:
        for f in _failed:
            print(f"  FAIL: {f}")
        return 1
    if args.keep:
        print(f"\nFiles written to {OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
