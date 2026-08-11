"""Excel generation for the tabular tasks (Day 4d).

Only Timeline, Find Entities and Compare Clauses have renderers; asking for any
other task raises, and the endpoint turns that into a 400 rather than handing
back a workbook containing a wall of prose in cell A1.

Every workbook carries a Metadata sheet with the attribution line and — because
it matters for trust — whether the data came from the structured intermediate
or from the markdown fallback.
"""

from __future__ import annotations

import datetime as dt
import io
import re
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .payload import ExportPayload
from .tables import ExportTable, tables_for, used_structured

HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, size=9, color="666666")
TITLE_FONT = Font(bold=True, size=12)
DRAFT_FONT = Font(bold=True, size=11, color="7A0000")

MAX_WIDTH = 62
MIN_WIDTH = 10
META_ROWS = 4  # 3 note lines + 1 blank before the header row
HEADER_ROW = META_ROWS + 1


class UnsupportedExport(RuntimeError):
    """Raised when a task has no Excel renderer."""


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel sheet names: <=31 chars, no []:*?/\\, and unique in the workbook."""
    clean = re.sub(r"[\[\]:*?/\\]", "-", (name or "Sheet").strip()) or "Sheet"
    clean = clean[:31]
    base, n = clean, 2
    while clean.lower() in used:
        suffix = f" ({n})"
        clean = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(clean.lower())
    return clean


def _write_meta_note(ws, payload: ExportPayload) -> None:
    """Three-line provenance note above the table. Deliberately in the sheet
    itself, not only on the Metadata tab: a lawyer who copies this sheet into
    another workbook takes the attribution with them."""
    ws.cell(1, 1, f"Matter: {payload.matter_name or '(ad-hoc query)'}").font = TITLE_FONT
    ws.cell(2, 1, payload.attribution()).font = NOTE_FONT
    ws.cell(
        3, 1, f"{payload.provenance_label}: " + ", ".join(payload.source_files)
    ).font = NOTE_FONT


def _autofit(ws, n_cols: int, first_row: int) -> None:
    for col in range(1, n_cols + 1):
        widest = 0
        for row in range(first_row, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            longest_line = max((len(seg) for seg in str(value).split("\n")), default=0)
            widest = max(widest, longest_line)
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(widest + 2, MIN_WIDTH), MAX_WIDTH
        )


def _write_table(
    ws, table: ExportTable, payload: ExportPayload, *, table_name: str
) -> None:
    """Write one ExportTable with a bold header row, wrapped cells, frozen
    header, and (where the headers are unique) a real Excel Table object."""
    _write_meta_note(ws, payload)
    n = len(table.headers)
    if not n:
        return

    for col, name in enumerate(table.headers, start=1):
        cell = ws.cell(HEADER_ROW, col, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for r, row in enumerate(table.rows, start=HEADER_ROW + 1):
        padded = (list(row) + [""] * n)[:n]
        for col, value in enumerate(padded, start=1):
            cell = ws.cell(r, col)
            idx = r - HEADER_ROW - 1
            if (
                table.date_column is not None
                and col == table.date_column + 1
                and idx in table.date_values
            ):
                # A real Excel date ONLY where the source date was complete and
                # unambiguous. Partial dates ("March 2024") stay verbatim text:
                # inventing a day to satisfy the column type would fabricate a
                # date in the column a lawyer is most likely to sort on.
                cell.value = dt.date.fromisoformat(table.date_values[idx])
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    last_row = HEADER_ROW + len(table.rows)
    ws.freeze_panes = ws.cell(HEADER_ROW + 1, 1)

    # An Excel Table requires unique, non-empty headers. Compare Clauses can
    # legitimately produce two columns with the same filename (the DB's
    # uniqueness is on content hash, not name — see ARCHITECTURE 2026-08-06), so
    # fall back to plain styled cells rather than emitting a workbook Excel
    # would refuse to open.
    unique = len({h.strip().lower() for h in table.headers}) == n and all(
        h.strip() for h in table.headers
    )
    if unique and table.rows:
        ref = f"A{HEADER_ROW}:{get_column_letter(n)}{last_row}"
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(t)

    if table.note:
        ws.cell(last_row + 2, 1, table.note).font = NOTE_FONT

    _autofit(ws, n, HEADER_ROW)


def _add_citations_sheet(wb: Workbook, payload: ExportPayload, used: set[str]) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Citations", used))
    ws.cell(1, 1, "Citations").font = TITLE_FONT
    ws.cell(2, 1, payload.attribution()).font = NOTE_FONT
    headers = ["#", "Source", "Locator", "Passage"]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(4, col, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, c in enumerate(payload.citations, start=1):
        ws.cell(4 + i, 1, i)
        ws.cell(4 + i, 2, c.source)
        ws.cell(4 + i, 3, c.page_or_paragraph)
        cell = ws.cell(4 + i, 4, c.text_snippet)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws.cell(5, 1)
    _autofit(ws, len(headers), 4)


def _add_metadata_sheet(wb: Workbook, payload: ExportPayload, used: set[str]) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Metadata", used))
    rows: list[tuple[str, str]] = [
        ("Task", payload.task_label),
        ("Matter", payload.matter_name or "(ad-hoc query)"),
        ("Generated", payload.attribution()),
        ("Model", payload.model),
        ("Embedding model", payload.embed_model),
        ("Retrieval depth", f"top-{payload.top_k}"),
        ("Run timestamp", payload.timestamp),
        (payload.provenance_label, ", ".join(payload.source_files)),
        (
            "Export data source",
            "structured task output"
            if used_structured(payload)
            else "parsed from the answer's markdown table (fallback)",
        ),
    ]
    if payload.task == "compare_clauses" and payload.comparison_table:
        # Which documents the model marked as lacking the clause is a FINDING,
        # not a gap — a lawyer needs to see that a file was checked.
        absent = _absent_files(payload)
        rows.append(
            (
                "Marked not present",
                ", ".join(absent) if absent else "(none - clause located in every file)",
            )
        )
    if payload.export_warnings:
        rows.append(("Warnings", "  |  ".join(payload.export_warnings)))
    if payload.is_pleading:
        rows.insert(0, ("DRAFT", "NOT FOR FILING OR SERVICE - NOT REVIEWED BY COUNSEL"))

    ws.cell(1, 1, "Export metadata").font = TITLE_FONT
    for r, (key, value) in enumerate(rows, start=3):
        ws.cell(r, 1, key).font = HEADER_FONT
        cell = ws.cell(r, 2, value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if key == "DRAFT":
            ws.cell(r, 1).font = DRAFT_FONT
            cell.font = DRAFT_FONT
    _autofit(ws, 2, 1)


def _absent_files(payload: ExportPayload) -> list[str]:
    """Files whose entire column says the clause is not present."""
    ct = payload.comparison_table
    if not ct or not ct.rows:
        return []
    absent: list[str] = []
    for i, name in enumerate(ct.files):
        values = [
            (r.cells[i] if i < len(r.cells) else "").strip().lower() for r in ct.rows
        ]
        if values and all(
            v.startswith("not present") or v in ("", "-", "—") for v in values
        ):
            absent.append(name)
    return absent


# --------------------------------------------------------------------------
# Per-task renderers
# --------------------------------------------------------------------------
def render_timeline(wb: Workbook, payload: ExportPayload, used: set[str]) -> None:
    tables = tables_for(payload)
    ws = wb.create_sheet(_safe_sheet_title("Timeline", used))
    if tables:
        _write_table(ws, tables[0], payload, table_name="TimelineTable")
    else:
        _write_meta_note(ws, payload)
        ws.cell(HEADER_ROW, 1, "No timeline table was produced for this result.")


def render_entities(wb: Workbook, payload: ExportPayload, used: set[str]) -> None:
    """One sheet per entity category — the shape a lawyer actually filters on,
    and it costs nothing because the categories are already separated."""
    tables = tables_for(payload)
    if not tables:
        ws = wb.create_sheet(_safe_sheet_title("Entities", used))
        _write_meta_note(ws, payload)
        ws.cell(HEADER_ROW, 1, "No entity tables were produced for this result.")
        return
    for i, table in enumerate(tables):
        ws = wb.create_sheet(_safe_sheet_title(table.title or f"Category {i+1}", used))
        if table.rows:
            _write_table(ws, table, payload, table_name=f"EntityTable{i+1}")
        else:
            _write_meta_note(ws, payload)
            ws.cell(
                HEADER_ROW,
                1,
                table.note or "None found in the retrieved passages.",
            ).font = NOTE_FONT


def render_comparison(wb: Workbook, payload: ExportPayload, used: set[str]) -> None:
    """Attributes down, documents across — the 2D shape the task produces.
    Citations stay as inline cell text rather than Excel comments, because
    comments do not survive a Google Sheets import and the test plan requires
    these files to open there."""
    tables = tables_for(payload)
    ws = wb.create_sheet(_safe_sheet_title("Comparison", used))
    if tables:
        _write_table(ws, tables[0], payload, table_name="ComparisonTable")
    else:
        _write_meta_note(ws, payload)
        ws.cell(HEADER_ROW, 1, "No comparison table was produced for this result.")


Renderer = Callable[[Workbook, ExportPayload, set], None]

XLSX_RENDERERS: dict[str, Renderer] = {
    "timeline": render_timeline,
    "find_entities": render_entities,
    "compare_clauses": render_comparison,
}


def build_xlsx(payload: ExportPayload) -> bytes:
    renderer = XLSX_RENDERERS.get(payload.task)
    if renderer is None:
        raise UnsupportedExport(
            f"{payload.task_label} produces prose, not a table — "
            f"export it as Word or PDF instead."
        )

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; renderers create their own
    used: set[str] = set()

    renderer(wb, payload, used)
    if payload.citations:
        _add_citations_sheet(wb, payload, used)
    _add_metadata_sheet(wb, payload, used)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
