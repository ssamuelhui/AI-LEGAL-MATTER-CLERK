r"""Excel (.xlsx) ingestion (Session 9).

Produces the same `Chunk(source, locator, text)` shape as every other format.

DESIGN CALIBRATED AGAINST THE PILOT LAWYER'S OWN SPREADSHEETS
-------------------------------------------------------------
Two real files drove every rule here, and each corrected an assumption:

  Submission List - Invoice   521 rows x 15 cols   76.8 tokens/row
  Invoice on May 20, 2026     379 rows x 82 cols   63.5 tokens/row

1. CHUNK BY TOKENS, NOT BY ROW COUNT. A fixed 50 rows would have produced
   3,176-3,839 token chunks against the 700-token target every other format
   uses. Oversized chunks distort cosine ranking against normal chunks, eat the
   top-k context budget (one such chunk exceeds five PDF chunks), and break the
   ~475-token-per-chunk assumption in Session 8's batch planner. Packing to
   ~700 tokens lands at 9-11 rows for these files.

2. THE HEADER IS NOT ALWAYS ROW 1. In `Invoice on May 20, 2026`, rows 1-2 are a
   title and a description and the real header is row 3. Assuming row 1 would
   have stamped "A. LIST OF STUDENTS..." onto every chunk of that sheet as a
   column list. Rows are scored instead; on the two real files the header wins
   by a clear margin (56 vs 43, and 72 vs 56 vs 0).

3. DROP EMPTY COLUMNS. That same sheet reports 82 columns and populates 18.
   Rendering all 82 makes every row two-thirds empty delimiters.

4. SKIP EMPTY ROWS, DO NOT SPLIT ON THEM. Every empty-row run in both files is
   a single row, and singles appear inside the data. A "5+ consecutive rows is
   a boundary" rule would never once have fired.

THE HEADER LINE IS REPEATED IN EVERY CHUNK, and that is not only for retrieval
quality: exhaustive Timeline has to know which column holds dates. A chunk of
bare values has no way to tell the model that column 4 is a date of birth
rather than an invoice date, so downstream extraction fails on unlabelled data.
"""

from __future__ import annotations

import datetime as dt
import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from .ingest import CHUNK_TARGET_TOKENS, Chunk

log = logging.getLogger("matter_clerk.ingest_xlsx")

# How far down to look for a header row before giving up.
HEADER_SCAN_ROWS = 10

# Columns quoted in the citation locator. Beyond this the citation is longer
# than the passage it labels.
LOCATOR_MAX_COLUMNS = 5


class XlsxUnreadable(RuntimeError):
    """The file could not be opened as a workbook."""


class XlsxPasswordProtected(XlsxUnreadable):
    """The file is encrypted."""


@dataclass
class SheetData:
    name: str
    header_row: int | None
    header_cells: list[str]
    rows: list[tuple[int, str]] = field(default_factory=list)   # (excel_row_no, text)
    has_formulas: bool = False
    formulas_without_values: bool = False


def _fmt(value) -> str:
    """Render a cell as a lawyer would expect to read it.

    Dates matter here: openpyxl hands back datetime objects, and str() on one
    gives "2025-01-31 00:00:00". A midnight timestamp on every date column is
    noise in the index and misleading in a citation, so a date-only value is
    rendered as a plain ISO date.
    """
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_score(cells: list[str]) -> int:
    """How much a row looks like a header rather than data or a title.

    Rewards breadth (a header fills most columns), short labels, and
    non-numeric content. A title row fills two cells and scores near zero; a
    data row is as wide as the header but carries numbers and longer values.
    """
    values = [c for c in cells if c]
    if len(values) < 3:
        return 0
    short = sum(1 for v in values if len(v) <= 28)
    non_numeric = sum(
        1 for v in values
        if not v.replace(".", "").replace("-", "").replace(",", "").isdigit()
    )
    return len(values) * 2 + short + non_numeric


def open_workbook(path: Path):
    """Open a workbook, distinguishing encryption from corruption.

    openpyxl raises InvalidFileException for both. An encrypted OOXML file is
    an OLE2 compound document with a fixed 8-byte signature; a healthy one is a
    zip starting "PK". Sniffing the header separates them without adding a
    dependency.
    """
    try:
        head = path.open("rb").read(8)
    except OSError as e:
        raise XlsxUnreadable(f"Could not read {path.name}: {e}") from e

    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        raise XlsxPasswordProtected(
            f"{path.name} is password protected. Remove the password in Excel, "
            "save a copy, and upload that copy."
        )
    try:
        # data_only: the cached computed value, not the formula text. A lawyer
        # cares what the cell says, not how it was worked out.
        return openpyxl.load_workbook(str(path), data_only=True, read_only=False)
    except Exception as e:                                        # noqa: BLE001
        raise XlsxUnreadable(
            f"{path.name} could not be opened as an Excel workbook. It may be "
            "damaged, or saved in an older .xls format -- re-save it as .xlsx."
        ) from e


def _has_formulas(path: Path) -> bool:
    """Whether any sheet contains a formula. Audit metadata only.

    Requires a second load without data_only, because one pass cannot see both
    the formula and its cached value.
    """
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False, read_only=True)
    except Exception:                                             # noqa: BLE001
        return False
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        return True
    except Exception:                                             # noqa: BLE001
        return False
    finally:
        wb.close()
    return False


def extract_xlsx(path: Path) -> list[SheetData]:
    """One SheetData per sheet, with the header located and empties dropped."""
    workbook = open_workbook(path)
    file_has_formulas = _has_formulas(path)
    sheets: list[SheetData] = []

    for worksheet in workbook.worksheets:
        raw = [list(r) for r in worksheet.iter_rows(values_only=True)]
        if not raw:
            continue

        # Which columns carry anything at all across the whole sheet.
        width = max((len(r) for r in raw), default=0)
        used = [
            i for i in range(width)
            if any(i < len(r) and _fmt(r[i]) for r in raw)
        ]
        if not used:
            sheets.append(SheetData(
                name=worksheet.title, header_row=None, header_cells=[],
                has_formulas=file_has_formulas,
                formulas_without_values=file_has_formulas,
            ))
            continue

        def cells_of(row) -> list[str]:
            return [_fmt(row[i]) if i < len(row) else "" for i in used]

        # Score the first rows and take the best as the header.
        best_row, best_score = None, 0
        for idx in range(min(HEADER_SCAN_ROWS, len(raw))):
            score = _header_score(cells_of(raw[idx]))
            if score > best_score:
                best_row, best_score = idx, score

        header_cells = cells_of(raw[best_row]) if best_row is not None else []
        header_excel_row = (best_row + 1) if best_row is not None else None

        rows: list[tuple[int, str]] = []
        for idx, row in enumerate(raw):
            if best_row is not None and idx <= best_row:
                continue                      # title rows and the header itself
            cells = cells_of(row)
            if not any(cells):
                continue                      # empty row: skipped, not a boundary
            rows.append((idx + 1, " | ".join(cells).strip()))

        sheets.append(SheetData(
            name=worksheet.title,
            header_row=header_excel_row,
            header_cells=header_cells,
            rows=rows,
            has_formulas=file_has_formulas,
            formulas_without_values=bool(file_has_formulas and not rows),
        ))

    workbook.close()
    return sheets


def _locator(source_sheet: str, first: int, last: int, header: list[str]) -> str:
    """`sheet 'Name', rows 15-24 (cols: A | B | C)`.

    Row numbers are real 1-based Excel rows so a lawyer can open the file and
    go straight there. The column list is what makes the row range mean
    something -- "rows 15-24" alone tells a reader nothing about what is in
    them -- truncated so a wide sheet does not produce a citation longer than
    the passage.
    """
    span = f"row {first}" if last <= first else f"rows {first}-{last}"
    base = f"sheet '{source_sheet}', {span}"
    named = [c for c in header if c][:LOCATOR_MAX_COLUMNS]
    if not named:
        return base
    more = "" if len([c for c in header if c]) <= LOCATOR_MAX_COLUMNS else " ..."
    return f"{base} (cols: {' | '.join(named)}{more})"


def chunk_xlsx(
    sheets: list[SheetData],
    source: str,
    chunk_tokens: int = CHUNK_TARGET_TOKENS,
) -> list[Chunk]:
    """Pack rows into ~700-token chunks, per sheet, header repeated on each.

    A chunk never spans two sheets: sheets are separate subjects, and a chunk
    straddling them would cite a row range that exists in neither.
    """
    from .ingest import _ENC

    def count(text: str) -> int:
        return len(_ENC.encode(text or ""))

    chunks: list[Chunk] = []
    for sheet in sheets:
        if not sheet.rows:
            continue
        header_line = " | ".join(sheet.header_cells).strip(" |")
        header_cost = count(header_line) if header_line else 0

        packed: list[str] = []
        first_row = sheet.rows[0][0]
        last_row = first_row
        running = header_cost

        for excel_row, text in sheet.rows:
            size = count(text)
            if packed and running + size > chunk_tokens:
                chunks.append(Chunk(
                    source=source,
                    locator=_locator(sheet.name, first_row, last_row,
                                     sheet.header_cells),
                    text=(f"{header_line}\n" if header_line else "")
                         + "\n".join(packed),
                ))
                packed = []
                running = header_cost
                first_row = excel_row
            packed.append(text)
            running += size
            last_row = excel_row

        if packed:
            chunks.append(Chunk(
                source=source,
                locator=_locator(sheet.name, first_row, last_row,
                                 sheet.header_cells),
                text=(f"{header_line}\n" if header_line else "")
                     + "\n".join(packed),
            ))
    return chunks


def extract_and_chunk(path: Path, source: str) -> tuple[list[Chunk], dict]:
    """Full pipeline for one .xlsx. Returns (chunks, stats)."""
    sheets = extract_xlsx(path)
    chunks = chunk_xlsx(sheets, source=source)
    stats = {
        "sheets": len(sheets),
        "sheet_names": [s.name for s in sheets],
        "rows": sum(len(s.rows) for s in sheets),
        "header_rows": {s.name: s.header_row for s in sheets},
        "has_formulas": any(s.has_formulas for s in sheets),
        # Every sheet held formulas and none held a cached value: the workbook
        # was written by a program and never opened in Excel, so there is
        # nothing to index and the remedy is different from a corrupt file.
        "formulas_without_values": bool(sheets)
        and all(s.formulas_without_values for s in sheets),
    }
    return chunks, stats
